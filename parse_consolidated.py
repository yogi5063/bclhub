#!/usr/bin/env python3
"""
parse_consolidated.py — Reads the Consolidated FIP workbook (uploaded per month)
and produces server/data_cache.json matching the dashboard schema.

Source workbook expected at: <UPLOAD_DIR>/<period>/Consolidated.xlsx
where period is 'YYYY-MM' (e.g. '2026-03'). If multiple periods exist,
all are loaded.

Output: server/data_cache.json with parsed[<Territory>||<period>] records.
"""
import os, sys, io, json, re, glob
from datetime import datetime
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl

SCRIPT_DIR = Path(__file__).parent
UPLOAD_DIR = Path(os.environ.get('UPLOAD_DIR', 'D:/Perklabs-mis/Upload'))
SETTINGS_FILE = SCRIPT_DIR / 'server' / 'settings.json'
if SETTINGS_FILE.exists():
    try:
        s = json.loads(SETTINGS_FILE.read_text(encoding='utf-8'))
        if 'uploadDir' in s:
            UPLOAD_DIR = Path(s['uploadDir'])
    except Exception:
        pass

OUTPUT_FILE = SCRIPT_DIR / 'server' / 'data_cache.json'

# Territory metadata
TERRITORIES = {
    'India':       {'brand': 'Basmi', 'currency': 'INR', 'fx_to_myr': 0.04274},
    'Malaysia':    {'brand': 'Basmi', 'currency': 'MYR', 'fx_to_myr': 1.0},
    'Philippines': {'brand': 'Basmi', 'currency': 'PHP', 'fx_to_myr': 0.0744},
    'Thailand':    {'brand': 'Basmi', 'currency': 'THB', 'fx_to_myr': 0.1255},
    'Indonesia':   {'brand': 'Basmi', 'currency': 'IDR', 'fx_to_myr': 0.0002374},
    'Brazil':      {'brand': 'Cure',  'currency': 'BRL', 'fx_to_myr': 0.752},
    'Brasil':      {'brand': 'Cure',  'currency': 'BRL', 'fx_to_myr': 0.752},
    'Europe':      {'brand': 'Cure',  'currency': 'EUR', 'fx_to_myr': 4.668},
    'GCC':         {'brand': 'Cure',  'currency': 'AED', 'fx_to_myr': 1.21},
    'Japan':       {'brand': 'Cure',  'currency': 'JPY', 'fx_to_myr': 0.02546},
    'Korea':       {'brand': 'Cure',  'currency': 'KRW', 'fx_to_myr': 0.002654},
    'Latam':       {'brand': 'Cure',  'currency': 'USD', 'fx_to_myr': 4.033},
    'Oceania':     {'brand': 'Cure',  'currency': 'AUD', 'fx_to_myr': 2.658},
    'USA':         {'brand': 'Cure',  'currency': 'USD', 'fx_to_myr': 4.033},
    'Molnu':       {'brand': 'Molnu', 'currency': 'USD', 'fx_to_myr': 4.033},
}

# Gateway_Reco_All column map (1-indexed)
GR_COLS = {
    'region': 1, 'order_no': 2, 'date': 3, 'customer': 4, 'pay_status': 6,
    'pay_method': 7, 'qty': 10, 'shipping_myr': 27, 'tax_myr': 28,
    'gross_myr': 30, 'refund_myr': 31, 'net_myr': 32,
    'px_net_myr': 40,
    'st_gross_myr': 46, 'st_fee_myr': 47, 'st_net_myr': 48,
    'pp_gross_myr': 54, 'pp_fee_myr': 55, 'pp_net_myr': 56,
    'xd_gross_myr': 62, 'xd_fee_myr': 63, 'xd_net_myr': 64,
    'settle_gross_myr': 82, 'settle_mdr_myr': 84, 'settle_net_myr': 85,
    'payment_myr': 107, 'dbt_myr': 114,
}

def to_num(v):
    if v is None or v == '': return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip())
    except: return 0.0

def parse_period_from_path(path):
    """Extract 'YYYY-MM' from folder name like 'Mar-26' or '2026-03'."""
    name = Path(path).parent.name.lower()
    # Try YYYY-MM
    m = re.match(r'^(\d{4})-(\d{2})$', name)
    if m:
        return f'{m.group(1)}-{m.group(2)}'
    # Try Mon-YY (e.g. mar-26)
    months = {'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06',
              'jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'}
    m = re.match(r'^([a-z]{3})-?(\d{2,4})$', name)
    if m:
        mon = months.get(m.group(1))
        if mon:
            yr = m.group(2)
            if len(yr) == 2: yr = '20' + yr
            return f'{yr}-{mon}'
    return None

def parse_workbook(fp, period):
    """Parse one Consolidated workbook → list of TerritoryResult dicts."""
    print(f'  Loading {fp} (data_only=True)...')
    wb = openpyxl.load_workbook(fp, data_only=True)
    print(f'    sheets: {len(wb.sheetnames)}')

    # Aggregator: territory -> dict of metrics
    per_terr = {}
    for tname in TERRITORIES:
        per_terr[tname] = {
            'territory': tname,
            'brand': TERRITORIES[tname]['brand'],
            # All money fields below are in MYR (sourced from Consolidated workbook MYR helpers)
            # 'local_currency' preserves the territory's native ccy for display
            'currency': 'MYR',
            'local_currency': TERRITORIES[tname]['currency'],
            'fx_rate_to_myr': TERRITORIES[tname]['fx_to_myr'],
            # P&L
            'gross': 0.0, 'shipping': 0.0, 'tax': 0.0,
            'refund_auto': 0.0, 'refund_manual': 0.0, 'refund_total': 0.0,
            'discount': 0.0, 'chargeback': 0.0,
            'net': 0.0, 'orders': 0, 'aov': 0.0, 'margin_pct': 0.0,
            # Fees per channel
            'fee_payex': 0.0, 'fee_paypal': 0.0, 'fee_stripe': 0.0,
            'fee_xendit': 0.0, 'fee_tiktok': 0.0, 'fee_shopee': 0.0, 'fee_lazada': 0.0,
            'fee_total': 0.0,
            # Channel revenue (gross paid via each gateway, in MYR)
            'gw_payex': 0.0, 'gw_paypal_gross': 0.0, 'gw_paypal_net': 0.0,
            'gw_stripe_gross': 0.0, 'gw_stripe_net': 0.0,
            'gw_xendit_gross': 0.0, 'gw_xendit_net': 0.0,
            'gw_settlement_net': 0.0,
            'payment': 0.0, 'dbt': 0.0,
            # Counts/details
            'orders_paid': 0, 'orders_unpaid': 0, 'orders_refunded': 0,
            'products': {},   # SKU -> dict
            'payment_methods': {},
            'states': {},
            'daily': {},
            'platforms': [{'name': 'Wix', 'gross_myr': 0, 'orders': 0}],
            'ar': {'payex_gross_myr': 0, 'payex_fee_myr': 0,
                   'payex_net_myr': 0, 'bank_receipts_myr': 0, 'ar_balance_myr': 0},
            'warnings': [],
            'errors': [],
            '_source': 'consolidated',
        }

    # Pre-aggregate Bank_Receipts by Region
    bank_receipts_by_region = {}
    if 'Bank_Receipts_All' in wb.sheetnames:
        br = wb['Bank_Receipts_All']
        # Find Region + Credit_MYR cols (last 2 cols typically)
        region_col = None; credit_myr_col = None
        for c in range(1, br.max_column+1):
            h = str(br.cell(1, c).value or '').strip().lower()
            if h == 'region': region_col = c
            if 'credit' in h and 'myr' in h: credit_myr_col = c
        if region_col and credit_myr_col:
            for r in range(2, br.max_row+1):
                rg = br.cell(r, region_col).value
                amt = br.cell(r, credit_myr_col).value
                if rg and isinstance(amt, (int, float)):
                    if rg == 'Brazil': rg = 'Brasil'
                    bank_receipts_by_region[rg] = bank_receipts_by_region.get(rg, 0) + float(amt)
    print(f'    Bank receipts by region: {sum(bank_receipts_by_region.values()):,.2f} MYR across {len(bank_receipts_by_region)} regions')

    # Process Gateway_Reco_All
    gws = wb['Gateway_Reco_All']
    print(f'    Gateway_Reco_All: {gws.max_row} rows')
    for r in range(4, gws.max_row + 1):
        region = gws.cell(r, GR_COLS['region']).value
        if not region: continue
        # Standardize Brazil → Brasil (consolidated uses Brasil throughout)
        if region == 'Brazil':
            region = 'Brasil'
        if region in per_terr:
            t = per_terr[region]
        else:
            continue

        gross   = to_num(gws.cell(r, GR_COLS['gross_myr']).value)
        net     = to_num(gws.cell(r, GR_COLS['net_myr']).value)
        ref     = to_num(gws.cell(r, GR_COLS['refund_myr']).value)
        ship    = to_num(gws.cell(r, GR_COLS['shipping_myr']).value)
        tax     = to_num(gws.cell(r, GR_COLS['tax_myr']).value)
        pay     = to_num(gws.cell(r, GR_COLS['payment_myr']).value)
        dbt     = to_num(gws.cell(r, GR_COLS['dbt_myr']).value)

        px_n    = to_num(gws.cell(r, GR_COLS['px_net_myr']).value)
        st_g    = to_num(gws.cell(r, GR_COLS['st_gross_myr']).value)
        st_f    = to_num(gws.cell(r, GR_COLS['st_fee_myr']).value)
        st_n    = to_num(gws.cell(r, GR_COLS['st_net_myr']).value)
        pp_g    = to_num(gws.cell(r, GR_COLS['pp_gross_myr']).value)
        pp_f    = to_num(gws.cell(r, GR_COLS['pp_fee_myr']).value)
        pp_n    = to_num(gws.cell(r, GR_COLS['pp_net_myr']).value)
        xd_g    = to_num(gws.cell(r, GR_COLS['xd_gross_myr']).value)
        xd_f    = to_num(gws.cell(r, GR_COLS['xd_fee_myr']).value)
        xd_n    = to_num(gws.cell(r, GR_COLS['xd_net_myr']).value)
        ss_g    = to_num(gws.cell(r, GR_COLS['settle_gross_myr']).value)
        ss_m    = to_num(gws.cell(r, GR_COLS['settle_mdr_myr']).value)
        ss_n    = to_num(gws.cell(r, GR_COLS['settle_net_myr']).value)

        pay_status = gws.cell(r, GR_COLS['pay_status']).value
        pay_method = gws.cell(r, GR_COLS['pay_method']).value

        t['gross']    += gross
        t['net']      += net
        t['refund_auto'] += ref
        t['refund_total'] += ref
        t['shipping'] += ship
        t['tax']      += tax
        t['payment']  += pay
        t['dbt']      += dbt

        t['gw_payex']         += px_n
        t['gw_paypal_gross']  += pp_g
        t['gw_paypal_net']    += pp_n
        t['gw_stripe_gross']  += st_g
        t['gw_stripe_net']    += st_n
        t['gw_xendit_gross']  += xd_g
        t['gw_xendit_net']    += xd_n
        t['gw_settlement_net']+= ss_n

        # Fee = -ve in our representation; we sum the negative values
        t['fee_paypal'] += abs(pp_f)
        # Stripe Fee from accountant has known inflation issue; recompute = Gross-Net
        t['fee_stripe'] += max(0, st_g - st_n)
        t['fee_xendit'] += xd_f
        t['fee_payex']  += ss_m

        t['orders'] += 1
        if pay_status == 'Paid': t['orders_paid'] += 1
        elif pay_status in ('Refunded','Partially refunded'): t['orders_refunded'] += 1
        else: t['orders_unpaid'] += 1

        if pay_method:
            t['payment_methods'][pay_method] = t['payment_methods'].get(pay_method, 0) + 1

        # Daily aggregation by date (col 3)
        date_v = gws.cell(r, GR_COLS['date']).value
        date_str = None
        if hasattr(date_v, 'strftime'):
            date_str = date_v.strftime('%Y-%m-%d')
        elif isinstance(date_v, str) and len(date_v) >= 10:
            date_str = date_v[:10]
        if date_str:
            d = t['daily'].setdefault(date_str, {'gross_myr':0, 'net_myr':0, 'orders':0})
            d['gross_myr'] += gross
            d['net_myr']   += net
            d['orders']    += 1

    # Process SKU_PL_All — per-SKU detail
    if 'SKU_PL_All' in wb.sheetnames:
        sku_ws = wb['SKU_PL_All']
        for r in range(3, sku_ws.max_row + 1):
            region = sku_ws.cell(r, 1).value
            sku = sku_ws.cell(r, 2).value
            if not region or not sku or sku == 'Region': continue
            # Standardize Brazil → Brasil (consolidated workbook bucket)
            if region == 'Brazil':
                region = 'Brasil'
            target = per_terr.get(region)
            if not target: continue
            target['products'][sku] = {
                'sku': sku,
                'description': sku_ws.cell(r, 3).value or '',
                'qty': to_num(sku_ws.cell(r, 6).value),
                'gross_myr': to_num(sku_ws.cell(r, 7).value),
                'net_myr': to_num(sku_ws.cell(r, 9).value),
                'cogs_myr': to_num(sku_ws.cell(r, 11).value),
                'gp_myr':  to_num(sku_ws.cell(r, 12).value),
            }

    # Compute derived
    results = []
    for tname, t in per_terr.items():
        # Skip empty territories
        if t['orders'] == 0 and t['gross'] == 0:
            continue
        # Compute fee_total
        t['fee_total'] = t['fee_paypal'] + t['fee_stripe'] + t['fee_xendit'] + t['fee_payex']
        # AOV
        t['aov'] = t['gross'] / max(t['orders'], 1)
        # Margin %
        t['margin_pct'] = (t['net'] / t['gross'] * 100) if t['gross'] else 0
        # Convert products dict to list
        t['products'] = list(t['products'].values())
        # Convert payment_methods dict to list
        t['payment_methods'] = [{'name': k, 'count': v} for k, v in t['payment_methods'].items()]
        # States — placeholder; daily already populated by main loop above
        t['states'] = []
        # Platforms: just Wix for now
        t['platforms'] = [{'name': 'Wix', 'gross_myr': t['gross'], 'orders': t['orders']}]
        # AR — augmented with bank receipts (newly tagged by region)
        bank_recv = bank_receipts_by_region.get(tname, 0.0)
        t['bank_receipts_myr'] = bank_recv
        t['ar'] = {
            'payex_gross_myr': t['gw_payex'],
            'payex_fee_myr': t['fee_payex'],
            'payex_net_myr': t['gw_settlement_net'],
            'bank_receipts_myr': bank_recv,
            'dbt_myr': t['dbt'],
            'expected_landed_myr': t['gw_settlement_net'] + t['dbt'],
            'ar_balance_myr': t['gw_settlement_net'] - bank_recv,
        }
        # Add to results with key {Territory}||{period}
        key = f'{tname}||{period}'
        results.append((key, t))

    wb.close()
    return results

def main():
    print(f'parse_consolidated.py — UPLOAD_DIR={UPLOAD_DIR}')
    if not UPLOAD_DIR.exists():
        print(f'!! UPLOAD_DIR not found: {UPLOAD_DIR}')
        sys.exit(1)

    # Find all Consolidated*.xlsx files under UPLOAD_DIR/<period>/
    candidates = []
    for fp in UPLOAD_DIR.glob('*/Consolidated*.xlsx'):
        period = parse_period_from_path(fp)
        if period:
            candidates.append((fp, period))
    # Also direct path UPLOAD_DIR/Consolidated*.xlsx → assume current/most-recent month
    for fp in UPLOAD_DIR.glob('Consolidated*.xlsx'):
        candidates.append((fp, '2026-03'))   # default if no period folder

    if not candidates:
        print(f'!! No Consolidated*.xlsx found under {UPLOAD_DIR}')
        sys.exit(1)

    print(f'Found {len(candidates)} workbook(s):')
    for fp, period in candidates:
        print(f'  {period} -> {fp}')

    all_results = {}
    for fp, period in candidates:
        try:
            for key, record in parse_workbook(fp, period):
                all_results[key] = record
            print(f'  ✓ Parsed {fp.name}')
        except Exception as e:
            import traceback
            print(f'  ✗ ERROR parsing {fp.name}: {e}')
            traceback.print_exc()

    output = {
        'generated_at': datetime.now().isoformat(),
        'parsed': all_results,
    }
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=None, ensure_ascii=False, default=str)
    print(f'\nWrote {OUTPUT_FILE}  ({len(all_results)} territory-period records)')

if __name__ == '__main__':
    main()
