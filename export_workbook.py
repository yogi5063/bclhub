#!/usr/bin/env python3
"""
export_workbook.py — FIP MIS Excel Report Generator
Reads data_cache.json (already parsed by parse_data.py) and generates a
professional Excel workbook with P&L, Reconciliation, and Territory breakdown.

Usage:
  python export_workbook.py                    # default: all periods
  python export_workbook.py --period 2026-01   # specific month
  python export_workbook.py --output path.xlsx # custom output path

Output: server/exports/FIP_MIS_Report_<period>.xlsx
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).parent
DATA_CACHE  = SCRIPT_DIR / 'server' / 'data_cache.json'
EXPORT_DIR  = SCRIPT_DIR / 'server' / 'exports'

# ── FX Rates (MYR base) ──────────────────────────────────────────────────────
FX_TO_MYR = {
    'MYR': 1.0, 'USD': 4.47, 'EUR': 4.84, 'GBP': 5.64, 'AUD': 2.88,
    'INR': 0.053, 'PHP': 0.078, 'THB': 0.131, 'IDR': 0.000272,
    'BRL': 0.79, 'AED': 1.217, 'JPY': 0.03, 'KRW': 0.00326,
    'SGD': 3.36, 'VND': 0.000178,
}

def to_myr(amount, currency):
    return float(amount) * FX_TO_MYR.get(str(currency).upper(), 1.0)

# ── Styles ────────────────────────────────────────────────────────────────────
THIN = Side(style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F_TITLE  = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
F_HDR    = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
F_BOLD   = Font(name='Calibri', size=10, bold=True)
F_NORM   = Font(name='Calibri', size=10)
F_TOTAL  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
F_NOTE   = Font(name='Calibri', size=9, italic=True, color='666666')

FILL_TITLE = PatternFill(fgColor='1F4E79', fill_type='solid')
FILL_HDR   = PatternFill(fgColor='2E75B6', fill_type='solid')
FILL_TOTAL = PatternFill(fgColor='1F4E79', fill_type='solid')
FILL_ALT1  = PatternFill(fgColor='F2F7FB', fill_type='solid')
FILL_ALT2  = PatternFill(fgColor='FFFFFF', fill_type='solid')
FILL_GREEN = PatternFill(fgColor='E2EFDA', fill_type='solid')
FILL_RED   = PatternFill(fgColor='FCE4EC', fill_type='solid')
FILL_AMBER = PatternFill(fgColor='FFF8E1', fill_type='solid')

AL = Alignment(horizontal='left', vertical='center')
AR = Alignment(horizontal='right', vertical='center')
AC = Alignment(horizontal='center', vertical='center')

def W(ws, r, c, val, font=F_NORM, fill=None, align=AL, fmt=None):
    cell = ws.cell(r, c)
    cell.value = val
    cell.font = font
    cell.alignment = align
    cell.border = BORDER
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    return cell

def title_row(ws, r, text, ncols, fill=FILL_TITLE):
    # Write value BEFORE merging
    cell = ws.cell(r, 1)
    cell.value = text; cell.font = F_TITLE; cell.fill = fill; cell.alignment = AL; cell.border = BORDER
    for c in range(2, ncols+1):
        cc = ws.cell(r, c); cc.fill = fill; cc.border = BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.row_dimensions[r].height = 28

def hdr_row(ws, r, labels, fill=FILL_HDR):
    for c, lbl in enumerate(labels, 1):
        W(ws, r, c, lbl, font=F_HDR, fill=fill, align=AC)
    ws.row_dimensions[r].height = 22

def data_fill(r):
    return FILL_ALT1 if r % 2 == 0 else FILL_ALT2


# ── Sheet Builders ────────────────────────────────────────────────────────────

def build_overview(wb, results, period_label):
    """KPI Overview sheet."""
    ws = wb.create_sheet("Overview")
    ncols = 6
    title_row(ws, 1, f"FIP MIS — Overview  |  {period_label}", ncols)

    hdr_row(ws, 3, ["Territory", "Brand", "Currency", "Gross Revenue",
                     "Net Revenue", "Net (MYR)"])

    r = 4
    total_gross_myr = 0
    total_net_myr = 0
    for res in sorted(results, key=lambda x: -to_myr(x['net'], x['currency'])):
        fill = data_fill(r)
        net_myr = to_myr(res['net'], res['currency'])
        gross_myr = to_myr(res['gross'], res['currency'])
        total_net_myr += net_myr
        total_gross_myr += gross_myr

        W(ws, r, 1, res['territory'], font=F_BOLD, fill=fill)
        W(ws, r, 2, res['brand'], fill=fill)
        W(ws, r, 3, res['currency'], fill=fill, align=AC)
        W(ws, r, 4, res['gross'], fill=fill, align=AR, fmt='#,##0.00')
        W(ws, r, 5, res['net'], fill=fill, align=AR, fmt='#,##0.00')
        W(ws, r, 6, net_myr, fill=fill, align=AR, fmt='#,##0.00')
        r += 1

    # Total row
    W(ws, r, 1, "TOTAL", font=F_TOTAL, fill=FILL_TOTAL)
    for c in range(2, 5):
        W(ws, r, c, None, fill=FILL_TOTAL)
    W(ws, r, 5, None, fill=FILL_TOTAL)
    W(ws, r, 6, total_net_myr, font=F_TOTAL, fill=FILL_TOTAL, align=AR, fmt='#,##0.00')

    # Widths
    for c, w in enumerate([18, 10, 8, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


def build_pl_summary(wb, results, period_label):
    """P&L Summary sheet — aggregated across all territories."""
    ws = wb.create_sheet("PL_Summary")
    ncols = 4
    title_row(ws, 1, f"P&L Summary  |  {period_label}  |  All Territories (MYR)", ncols)

    hdr_row(ws, 3, ["P&L Line", "MYR Amount", "% of Gross", "Source"])

    total_gross = sum(to_myr(r['gross'], r['currency']) for r in results)
    total_ship = sum(to_myr(r['shipping'], r['currency']) for r in results)
    total_refund = sum(to_myr(r['refund_total'], r['currency']) for r in results)
    total_fees = sum(to_myr(r['fee_total'], r['currency']) for r in results)
    total_disc = sum(to_myr(r['discount'], r['currency']) for r in results)
    total_net = sum(to_myr(r['net'], r['currency']) for r in results)
    total_orders = sum(r['orders'] for r in results)

    lines = [
        ("Gross Revenue", total_gross, "SUM of all territory Gross × FX"),
        ("Less: Shipping", -total_ship, "SUM of Shipping across territories"),
        ("Less: Refunds", -total_refund, "SUM of all refunds (auto + manual + chargeback)"),
        ("Less: Discounts", -total_disc, "SUM of Discount column per territory"),
        ("Less: Gateway Fees", -total_fees, "Payex + PayPal + Xendit + Marketplace fees"),
        ("NET REVENUE", total_net, "Gross - Shipping - Refunds - Discounts - Fees"),
    ]

    r = 4
    for label, amount, source in lines:
        fill = data_fill(r)
        is_total = label == "NET REVENUE"
        _font = F_TOTAL if is_total else F_NORM
        _fill = FILL_TOTAL if is_total else fill
        pct = (amount / total_gross * 100) if total_gross > 0 else 0
        W(ws, r, 1, label, font=_font if not is_total else Font(name='Calibri', size=11, bold=True, color='FFFFFF'), fill=_fill)
        W(ws, r, 2, amount, font=_font if not is_total else Font(name='Calibri', size=11, bold=True, color='FFFFFF'), fill=_fill, align=AR, fmt='#,##0.00')
        W(ws, r, 3, f"{pct:.1f}%", font=_font if not is_total else Font(name='Calibri', size=11, bold=True, color='FFFFFF'), fill=_fill, align=AR)
        W(ws, r, 4, source, font=F_NOTE if not is_total else Font(name='Calibri', size=9, italic=True, color='FFFFFF'), fill=_fill)
        r += 1

    r += 1
    W(ws, r, 1, "KPIs", font=F_BOLD)
    r += 1
    W(ws, r, 1, "Total Orders", font=F_NORM)
    W(ws, r, 2, total_orders, align=AR, fmt='#,##0')
    r += 1
    aov = total_gross / total_orders if total_orders > 0 else 0
    W(ws, r, 1, "Average Order Value (MYR)", font=F_NORM)
    W(ws, r, 2, aov, align=AR, fmt='#,##0.00')
    r += 1
    margin = total_net / total_gross * 100 if total_gross > 0 else 0
    W(ws, r, 1, "Net Margin %", font=F_NORM)
    W(ws, r, 2, f"{margin:.1f}%", align=AR)

    for c, w in enumerate([28, 16, 12, 50], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


def build_territory_detail(wb, results, period_label):
    """Territory P&L breakdown — one row per territory."""
    ws = wb.create_sheet("Territory_Detail")
    cols = ["Territory", "Brand", "CCY", "Gross (Local)", "Shipping", "Refunds",
            "Discounts", "Fees", "Net (Local)", "FX Rate", "Gross (MYR)",
            "Net (MYR)", "Margin %", "Orders", "AOV (MYR)"]
    ncols = len(cols)
    title_row(ws, 1, f"Territory P&L Detail  |  {period_label}", ncols)
    hdr_row(ws, 3, cols)

    r = 4
    tot_gross_myr = 0
    tot_net_myr = 0
    tot_orders = 0
    for res in sorted(results, key=lambda x: -to_myr(x['net'], x['currency'])):
        fill = data_fill(r)
        ccy = res['currency']
        fx = FX_TO_MYR.get(ccy.upper(), 1.0) if ccy else 1.0
        g_myr = to_myr(res['gross'], ccy)
        n_myr = to_myr(res['net'], ccy)
        margin = (res['net'] / res['gross'] * 100) if res['gross'] > 0 else 0
        aov_myr = g_myr / res['orders'] if res['orders'] > 0 else 0

        tot_gross_myr += g_myr
        tot_net_myr += n_myr
        tot_orders += res['orders']

        vals = [res['territory'], res['brand'], ccy, res['gross'], res['shipping'],
                res['refund_total'], res['discount'], res['fee_total'], res['net'],
                fx, g_myr, n_myr, margin, res['orders'], aov_myr]
        fmts = [None, None, None, '#,##0.00', '#,##0.00', '#,##0.00', '#,##0.00',
                '#,##0.00', '#,##0.00', '0.00000', '#,##0.00', '#,##0.00',
                '0.0"%"', '#,##0', '#,##0.00']

        for ci, (v, fm) in enumerate(zip(vals, fmts), 1):
            _a = AR if ci >= 4 else AL
            W(ws, r, ci, v, fill=fill, align=_a, fmt=fm,
              font=F_BOLD if ci == 1 else F_NORM)
        r += 1

    # Total
    W(ws, r, 1, "TOTAL", font=F_TOTAL, fill=FILL_TOTAL)
    for c in range(2, ncols+1):
        W(ws, r, c, None, fill=FILL_TOTAL)
    W(ws, r, 11, tot_gross_myr, font=F_TOTAL, fill=FILL_TOTAL, align=AR, fmt='#,##0.00')
    W(ws, r, 12, tot_net_myr, font=F_TOTAL, fill=FILL_TOTAL, align=AR, fmt='#,##0.00')
    tot_margin = tot_net_myr / tot_gross_myr * 100 if tot_gross_myr > 0 else 0
    W(ws, r, 13, tot_margin, font=F_TOTAL, fill=FILL_TOTAL, align=AR, fmt='0.0"%"')
    W(ws, r, 14, tot_orders, font=F_TOTAL, fill=FILL_TOTAL, align=AR, fmt='#,##0')

    widths = [18, 8, 6, 14, 12, 12, 12, 12, 14, 10, 14, 14, 10, 10, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


def build_products(wb, results, period_label):
    """Top products across all territories."""
    ws = wb.create_sheet("Products")
    cols = ["Rank", "Product", "Territory", "Orders", "Revenue (Local)",
            "Revenue (MYR)", "Share %"]
    ncols = len(cols)
    title_row(ws, 1, f"Product Analysis  |  {period_label}  |  Top 100", ncols)
    hdr_row(ws, 3, cols)

    # Aggregate products across territories
    prod_map = {}
    for res in results:
        ccy = res['currency']
        for p in res.get('products', []):
            key = p['name']
            if key not in prod_map:
                prod_map[key] = {'name': key, 'territory': res['territory'],
                                 'orders': 0, 'revenue_local': 0, 'revenue_myr': 0}
            prod_map[key]['orders'] += p['orders']
            prod_map[key]['revenue_local'] += p['revenue']
            prod_map[key]['revenue_myr'] += to_myr(p['revenue'], ccy)

    total_myr = sum(v['revenue_myr'] for v in prod_map.values())
    sorted_prods = sorted(prod_map.values(), key=lambda x: -x['revenue_myr'])[:100]

    r = 4
    for rank, p in enumerate(sorted_prods, 1):
        fill = data_fill(r)
        share = p['revenue_myr'] / total_myr * 100 if total_myr > 0 else 0
        vals = [rank, p['name'], p['territory'], p['orders'],
                p['revenue_local'], p['revenue_myr'], share]
        fmts = ['#,##0', None, None, '#,##0', '#,##0.00', '#,##0.00', '0.0"%"']
        for ci, (v, fm) in enumerate(zip(vals, fmts), 1):
            W(ws, r, ci, v, fill=fill, align=AR if ci >= 4 else AL, fmt=fm)
        r += 1

    widths = [6, 40, 14, 10, 14, 14, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


def build_payments(wb, results, period_label):
    """Payment method breakdown."""
    ws = wb.create_sheet("Payments")
    cols = ["Territory", "Payment Method", "Orders", "Revenue (Local)",
            "Revenue (MYR)", "Share %"]
    ncols = len(cols)
    title_row(ws, 1, f"Payment Methods  |  {period_label}", ncols)
    hdr_row(ws, 3, cols)

    r = 4
    for res in sorted(results, key=lambda x: -to_myr(x['gross'], x['currency'])):
        ccy = res['currency']
        for pm in res.get('payment_methods', []):
            fill = data_fill(r)
            rev_myr = to_myr(pm['revenue'], ccy)
            share = pm['revenue'] / res['gross'] * 100 if res['gross'] > 0 else 0
            W(ws, r, 1, res['territory'], font=F_BOLD, fill=fill)
            W(ws, r, 2, pm['method'], fill=fill)
            W(ws, r, 3, pm['orders'], fill=fill, align=AR, fmt='#,##0')
            W(ws, r, 4, pm['revenue'], fill=fill, align=AR, fmt='#,##0.00')
            W(ws, r, 5, rev_myr, fill=fill, align=AR, fmt='#,##0.00')
            W(ws, r, 6, share, fill=fill, align=AR, fmt='0.0"%"')
            r += 1

    widths = [18, 22, 10, 16, 16, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


def build_daily(wb, results, period_label):
    """Daily revenue trend for the period."""
    ws = wb.create_sheet("Daily_Trend")
    cols = ["Date", "Orders", "Revenue (MYR)"]
    ncols = len(cols)
    title_row(ws, 1, f"Daily Revenue Trend  |  {period_label}  |  All Territories (MYR)", ncols)
    hdr_row(ws, 3, cols)

    # Aggregate daily across territories
    daily_map = {}
    for res in results:
        ccy = res['currency']
        for dk, v in res.get('daily', {}).items():
            if dk not in daily_map:
                daily_map[dk] = {'orders': 0, 'revenue_myr': 0}
            daily_map[dk]['orders'] += v['orders']
            daily_map[dk]['revenue_myr'] += to_myr(v['revenue'], ccy)

    r = 4
    for dk in sorted(daily_map.keys()):
        fill = data_fill(r)
        v = daily_map[dk]
        W(ws, r, 1, dk, fill=fill)
        W(ws, r, 2, v['orders'], fill=fill, align=AR, fmt='#,##0')
        W(ws, r, 3, v['revenue_myr'], fill=fill, align=AR, fmt='#,##0.00')
        r += 1

    for c, w in enumerate([14, 10, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


def build_recon(wb, results, period_label):
    """Reconciliation summary — gateway vs calculated."""
    ws = wb.create_sheet("Reconciliation")
    ncols = 5
    title_row(ws, 1, f"Reconciliation Summary  |  {period_label}  |  MYR", ncols)

    hdr_row(ws, 3, ["Check", "Expected", "Actual", "Difference", "Status"])

    total_gross = sum(to_myr(r['gross'], r['currency']) for r in results)
    total_net = sum(to_myr(r['net'], r['currency']) for r in results)
    total_fees = sum(to_myr(r['fee_total'], r['currency']) for r in results)
    total_ship = sum(to_myr(r['shipping'], r['currency']) for r in results)
    total_ref = sum(to_myr(r['refund_total'], r['currency']) for r in results)
    total_disc = sum(to_myr(r['discount'], r['currency']) for r in results)
    calc_net = total_gross - total_ship - total_ref - total_disc - total_fees

    checks = [
        ("Net = Gross - Ship - Ref - Disc - Fees", total_net, calc_net),
        ("Fees < 15% of Gross", total_gross * 0.15, total_fees),
        ("Refunds < 5% of Gross", total_gross * 0.05, total_ref),
    ]

    r = 4
    for label, expected, actual in checks:
        fill = data_fill(r)
        diff = actual - expected
        status = "\u2713 OK" if abs(diff) < max(abs(expected) * 0.01, 100) else "\u26a0 CHECK"
        s_fill = FILL_GREEN if "\u2713" in status else FILL_AMBER
        W(ws, r, 1, label, font=F_BOLD, fill=fill)
        W(ws, r, 2, expected, fill=fill, align=AR, fmt='#,##0.00')
        W(ws, r, 3, actual, fill=fill, align=AR, fmt='#,##0.00')
        W(ws, r, 4, diff, fill=fill, align=AR, fmt='#,##0.00')
        W(ws, r, 5, status, fill=s_fill, align=AC)
        r += 1

    # Summary KPIs
    r += 1
    W(ws, r, 1, "SUMMARY", font=F_BOLD, fill=FILL_HDR)
    W(ws, r, 2, None, fill=FILL_HDR); W(ws, r, 3, None, fill=FILL_HDR)
    r += 1
    W(ws, r, 1, "Total Gross (MYR)"); W(ws, r, 2, total_gross, align=AR, fmt='#,##0.00')
    r += 1
    W(ws, r, 1, "Total Net (MYR)"); W(ws, r, 2, total_net, align=AR, fmt='#,##0.00')
    r += 1
    W(ws, r, 1, "Total Fees (MYR)"); W(ws, r, 2, total_fees, align=AR, fmt='#,##0.00')
    r += 1
    W(ws, r, 1, "Total Orders"); W(ws, r, 2, sum(r_['orders'] for r_ in results), align=AR, fmt='#,##0')
    r += 1
    W(ws, r, 1, "Territories"); W(ws, r, 2, len(results), align=AR, fmt='#,##0')

    for c, w in enumerate([40, 16, 16, 16, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    return ws


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='FIP MIS Excel Report Generator')
    parser.add_argument('--period', type=str, default=None,
                        help='Period filter (e.g. 2026-01). Default: all data.')
    parser.add_argument('--output', type=str, default=None,
                        help='Output file path. Default: server/exports/FIP_MIS_Report_<period>.xlsx')
    args = parser.parse_args()

    # Load cached data
    if not DATA_CACHE.exists():
        print(f"ERROR: {DATA_CACHE} not found. Run parse_data.py first.")
        sys.exit(1)

    with open(DATA_CACHE, 'r', encoding='utf-8') as f:
        cache = json.load(f)

    parsed = cache.get('parsed', {})
    if not parsed:
        print("ERROR: No parsed data in cache.")
        sys.exit(1)

    # Filter results
    results = []
    for key, val in parsed.items():
        # Skip full-year aggregates (those have "||wix" keys)
        if '||wix' in key:
            continue
        # Apply period filter
        if args.period:
            parts = key.split('||')
            if len(parts) >= 2 and parts[1] != args.period:
                continue
        results.append(val)

    if not results:
        # Fallback: use full-year results
        results = [v for k, v in parsed.items() if '||wix' in k]

    if not results:
        print("ERROR: No results match the filter.")
        sys.exit(1)

    period_label = args.period or "All Periods"
    print(f"[export] Generating report for: {period_label}")
    print(f"[export] Territories: {len(results)}")

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    build_overview(wb, results, period_label)
    build_pl_summary(wb, results, period_label)
    build_territory_detail(wb, results, period_label)
    build_products(wb, results, period_label)
    build_payments(wb, results, period_label)
    build_daily(wb, results, period_label)
    build_recon(wb, results, period_label)

    # Save
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    if args.output:
        out_path = Path(args.output)
    else:
        safe_period = (args.period or 'all').replace('-', '_')
        out_path = EXPORT_DIR / f"FIP_MIS_Report_{safe_period}.xlsx"

    wb.save(str(out_path))
    print(f"[export] Saved: {out_path}")
    print(f"[export] Sheets: {len(wb.sheetnames)}")
    # Print path for Node.js to capture
    print(f"OUTPUT_PATH={out_path}")

if __name__ == '__main__':
    main()
