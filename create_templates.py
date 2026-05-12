"""
create_templates.py
Creates all 7 Excel upload templates with correct headers + sample rows.
Run once: python create_templates.py
Templates saved to: templates/ folder
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
os.makedirs(TEMPLATES_DIR, exist_ok=True)

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
SAMPLE_FILL  = PatternFill("solid", fgColor="EBF3FB")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
SAMPLE_FONT  = Font(color="555555", italic=True, size=9)
NOTE_FILL    = PatternFill("solid", fgColor="FFF3CD")
NOTE_FONT    = Font(color="856404", bold=True, size=9)

def style_header(ws, headers, col_widths=None):
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.row_dimensions[1].height = 30
        if col_widths and ci <= len(col_widths):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1]
        else:
            ws.column_dimensions[get_column_letter(ci)].width = 18

def add_sample_row(ws, row_data, row_num=2):
    for ci, v in enumerate(row_data, 1):
        cell = ws.cell(row_num, ci, value=v)
        cell.fill = SAMPLE_FILL
        cell.font = SAMPLE_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')

def add_note(ws, text, row, col=1, span=4):
    cell = ws.cell(row, col, value=f"ℹ️  {text}")
    cell.fill = NOTE_FILL
    cell.font = NOTE_FONT
    ws.row_dimensions[row].height = 20

# ── 1. Wix Orders ─────────────────────────────────────────────────────────────
def make_wix_orders():
    wb = Workbook(); ws = wb.active; ws.title = "DATA_Orders"
    headers = [
        "Region","Order Number","Date Created","Billing First Name","Billing Last Name",
        "Billing Email","Billing Phone","Billing Address","Billing City","Billing State",
        "Billing Country","Billing ZIP","Shipping First Name","Shipping Last Name",
        "Shipping Email","Shipping Phone","Shipping Address","Shipping City","Shipping State",
        "Shipping Country","Shipping ZIP","Buyer Note","Order Notes","Items","Item Count",
        "Weight (kg)","Delivery Method","Carrier","Tracking Number","Tracking Link",
        "Coupon Code","Coupon Discount","Subtotal","Discount","Tax","Shipping Rate",
        "Total","Refunded Amount","Net Amount","Currency","Payment Status","Fulfillment Status"
    ]
    widths = [12,18,16,14,14,22,14,24,14,14,12,10,14,14,22,14,24,14,14,12,10,
              18,18,30,10,10,16,14,18,30,16,14,14,14,12,14,14,14,14,10,18,18]
    style_header(ws, headers, widths)
    sample = [
        "Malaysia","WIX-123456","2026-04-01","John","Doe","john@email.com","0123456789",
        "123 Jalan Ampang","Kuala Lumpur","Selangor","Malaysia","50450",
        "John","Doe","john@email.com","0123456789","123 Jalan Ampang","Kuala Lumpur",
        "Selangor","Malaysia","50450","","",
        "CureFIP 1500mg x1","1","0.5","Standard","Poslaju","TR123456",
        "https://track.poslaju.com","SAVE10","10.00","290.00","10.00","0","10.00",
        "290.00","0","290.00","290.00","MYR","Paid","Fulfilled"
    ]
    add_sample_row(ws, sample)
    add_note(ws, "Row 1 = Headers (DO NOT change). Row 2 = Sample (delete before upload). Add your data from Row 3 onwards.", 3)
    wb.save(os.path.join(TEMPLATES_DIR, 'Wix_Orders_Template.xlsx'))
    print("✓ Wix_Orders_Template.xlsx")

# ── 2. Wix Payments ───────────────────────────────────────────────────────────
def make_wix_payments():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    headers = [
        "Payment ID","Created Date","Type","Merchant ID","Order ID",
        "Currency","Amount","Processing Fee","Service Fee","Net Amount",
        "Transaction Status","Transaction ID","Payment Provider","Payment Method",
        "Card Type","Card Last 4 Digits","Billing Name","Billing Email",
        "Refund ID","Refund Amount","Refund Reason","Subscription ID",
        "Subscription Cycle","Payout ID","Payout Date","Payout Amount",
        "Payout Currency","Payout Status","Gateway Reference","Notes",
        "Buyer Name","Buyer Email","Buyer Phone","Shipping Address","Country",
        "State","City","ZIP","Coupon Code","Discount Amount","Tax Amount",
        "Item Name","Item SKU","Item Price","Item Quantity","Item Total",
        "Shipping Total","Total Charges","Payment Source","Checkout ID",
        "Order Number","Product Type","Subscription Status","Member ID",
        "Plan Name","Transaction Type","Processor Name","Order ID (text)",
        "Order Source","Notes 2"
    ]
    style_header(ws, headers)
    sample = ["PAY-001","2026-04-01","Payment","MID-001","12345678",
              "MYR","290.00","","","290.00","Successful","TXN-001",
              "Payex","Credit Card","Visa","4242","John Doe","john@email.com",
              "","","","","","","","","","","","",
              "John Doe","john@email.com","0123456789","123 Jalan Ampang","Malaysia",
              "Selangor","KL","50450","","","0",
              "CureFIP 1500mg","SKU-001","290.00","1","290.00",
              "10.00","290.00","Online","CHK-001",
              "WIX-123456","Physical","","",
              "","Payment","Payex","12345678","Wix",""]
    add_sample_row(ws, sample)
    add_note(ws, "Sheet name must stay 'Sheet1'. Order ID in column AT (column 46). Currency in column E.", 3)
    wb.save(os.path.join(TEMPLATES_DIR, 'Wix_Payments_Template.xlsx'))
    print("✓ Wix_Payments_Template.xlsx")

# ── 3. Stripe Report ──────────────────────────────────────────────────────────
def make_stripe():
    wb = Workbook(); ws = wb.active; ws.title = "Stripe report 2"
    headers = ["id","Type","Source","Amount (MYR)","Fee (MYR)","Net (MYR)",
               "Currency","Description","Created (UTC)","Available On (UTC)",
               "Transfer","Transfer Date","Transfer Group","wix_transaction_id",
               "customer_email","customer_name","billing_country"]
    style_header(ws, headers)
    sample = ["txn_001","charge","ch_001","686.00","11.00","675.00",
              "myr","CureFIP Order","2026-04-01 10:00:00","2026-04-03 10:00:00",
              "po_001","2026-04-03","","abc-123-def-456",
              "john@email.com","John Doe","KR"]
    add_sample_row(ws, sample)
    add_note(ws, "Sheet name must be 'Stripe report 2'. Amount/Fee/Net already in MYR. Filter Type='charge' rows only.", 3)
    wb.save(os.path.join(TEMPLATES_DIR, 'Stripe_Template.xlsx'))
    print("✓ Stripe_Template.xlsx")

# ── 4. PayPal Report ──────────────────────────────────────────────────────────
def make_paypal():
    wb = Workbook(); ws = wb.active; ws.title = "Paypal"
    headers = ["Date","Time","Timezone","Description","Currency","Gross",
               "Fee","Net","From Email Address","Name","Transaction ID",
               "Receipt ID","Balance","Address Line 1","Address Line 2",
               "City","State","ZIP","Country","Contact Phone","Subject",
               "Note","Country Code","Balance Impact",
               "Custom Number","Invoice Number","Reference TX ID"]
    style_header(ws, headers)
    sample = ["01/04/2026","10:00:00","GMT+8","Express Checkout Payment",
              "USD","130.00","-3.90","126.10","buyer@email.com","John Doe",
              "PAYID-001","RCP-001","1000.00","123 Main St","",
              "Seoul","","","KR","","CureFIP Order","",
              "KR","Debit","wix-txn-guid-here","INV-001",""]
    add_sample_row(ws, sample)
    add_note(ws, "Sheet name must be 'Paypal'. Filter: Description='Express Checkout Payment' AND Balance Impact='Debit'. Fee is negative.", 3)
    wb.save(os.path.join(TEMPLATES_DIR, 'PayPal_Template.xlsx'))
    print("✓ PayPal_Template.xlsx")

# ── 5. Payex Settlement Report ────────────────────────────────────────────────
def make_payex():
    wb = Workbook(); ws = wb.active; ws.title = "Payex Report 2"
    headers = [
        "Merchant ID","Merchant Name","Settlement Date","Settlement Period Start",
        "Settlement Period End","Settlement ID","Transaction ID","Order Reference",
        "Transaction Type","Transaction Date","Transaction Time","Currency",
        "Transaction Amount","Settlement Amount (MYR)","Exchange Rate",
        "BaseMDR","MDR Amount (MYR)","Processing Fee","Service Tax","Total Deduction",
        "Net Settlement (MYR)","Payment Method","Card Type","Card Issuer",
        "wixTransactionId","Customer Name","Customer Email"
    ]
    style_header(ws, headers)
    sample = ["MID-001","Citia Trading","2026-04-03","2026-04-01","2026-04-03",
              "SETTLE-001","TXN-001","WIX-123456","Settlement","2026-04-01","10:00:00",
              "MYR","290.00","290.00","1.0",
              "1.50%","4.35","0","0","4.35",
              "285.65","Credit Card","Visa","Maybank",
              "wix-txn-guid-here","John Doe","john@email.com"]
    add_sample_row(ws, sample)
    add_note(ws, "Sheet name must be 'Payex Report 2'. MDR Amount in MYR already. Filter Type='Settlement'.", 3)
    wb.save(os.path.join(TEMPLATES_DIR, 'Payex_Template.xlsx'))
    print("✓ Payex_Template.xlsx")

# ── 6. Marketplace (TikTok/Shopee/Lazada) ────────────────────────────────────
def make_marketplace():
    wb = Workbook()

    # TikTok
    ws1 = wb.active; ws1.title = "TikTok"
    h1 = ["Order ID","Order Status","SKU","Product Name","Quantity","Selling Price",
          "Subtotal","Shipping Fee Charged","Total Fees","Net Revenue",
          "Currency","Settlement Date","Region"]
    style_header(ws1, h1)
    add_sample_row(ws1, ["TT-001","Completed","SKU-001","CureFIP 1500mg","1",
                         "290.00","290.00","0","14.50","275.50","MYR","2026-04-03","Malaysia"])
    add_note(ws1, "TikTok orders. 'Total Fees' = platform commission.", 3)

    # Shopee
    ws2 = wb.create_sheet("Shopee")
    h2 = ["Order ID","Order Status","SKU","Product Name","Quantity","Selling Price",
          "Subtotal","Commission","Service Fee","Transaction Fee",
          "Total Fees","Net Revenue","Currency","Settlement Date","Region"]
    style_header(ws2, h2)
    add_sample_row(ws2, ["SP-001","Completed","SKU-001","CureFIP 1500mg","1",
                         "290.00","290.00","8.70","1.45","0.87",
                         "11.02","278.98","MYR","2026-04-03","Malaysia"])
    add_note(ws2, "Shopee orders. Commission + Service Fee + Transaction Fee = Total Fees.", 3)

    # Lazada
    ws3 = wb.create_sheet("Lazada")
    h3 = ["Order ID","Order Status","SKU","Product Name","Quantity","Unit Price",
          "Subtotal","Commission Rate","Commission Amount","Shipping Fee","Total Fees",
          "Net Revenue","Currency","Settlement Date","Region"]
    style_header(ws3, h3)
    add_sample_row(ws3, ["LZ-001","Delivered","SKU-001","CureFIP 1500mg","1",
                         "290.00","290.00","5%","14.50","0","14.50",
                         "275.50","MYR","2026-04-03","Malaysia"])
    add_note(ws3, "Lazada orders. Commission Amount = platform fee.", 3)

    wb.save(os.path.join(TEMPLATES_DIR, 'Marketplace_Template.xlsx'))
    print("✓ Marketplace_Template.xlsx (TikTok + Shopee + Lazada)")

# ── 7. FX Rates ───────────────────────────────────────────────────────────────
def make_fx():
    wb = Workbook(); ws = wb.active; ws.title = "FX_Rates"
    headers = ["Currency Code","Currency Name","Rate to MYR","Source","Notes"]
    style_header(ws, headers, [15, 22, 15, 20, 30])
    rates = [
        ("KRW","Korean Won","0.002654","Bank Negara",""),
        ("JPY","Japanese Yen","0.02522","Bank Negara",""),
        ("EUR","Euro","4.619","Bank Negara",""),
        ("USD","US Dollar","4.447","Bank Negara",""),
        ("AED","UAE Dirham","1.098","Estimate",""),
        ("MXN","Mexican Peso","0.2215","Estimate",""),
        ("BRL","Brazilian Real","0.7675","Estimate",""),
        ("AUD","Australian Dollar","2.759","Bank Negara",""),
        ("INR","Indian Rupee","0.04274","Estimate",""),
        ("IDR","Indonesian Rupiah","0.0002374","Bank Negara",""),
        ("PHP","Philippine Peso","0.07748","Estimate",""),
        ("THB","Thai Baht","0.1227","Estimate",""),
        ("MYR","Malaysian Ringgit","1.0","Base Currency",""),
        ("SGD","Singapore Dollar","3.29","Bank Negara",""),
        ("GBP","British Pound","5.62","Bank Negara",""),
        ("HKD","Hong Kong Dollar","0.571","Bank Negara",""),
        ("CNY","Chinese Yuan","0.611","Bank Negara",""),
        ("TWD","Taiwan Dollar","0.136","Bank Negara",""),
        ("CHF","Swiss Franc","4.98","Bank Negara",""),
    ]
    for i, r in enumerate(rates, 2):
        for j, v in enumerate(r, 1):
            ws.cell(i, j, value=v)
    add_note(ws, "Update rates monthly. MYR always = 1.0. Check Bank Negara for official rates.", len(rates)+3)
    wb.save(os.path.join(TEMPLATES_DIR, 'FX_Rates_Template.xlsx'))
    print("✓ FX_Rates_Template.xlsx")

if __name__ == '__main__':
    print(f"Creating templates in: {TEMPLATES_DIR}\n")
    make_wix_orders()
    make_wix_payments()
    make_stripe()
    make_paypal()
    make_payex()
    make_marketplace()
    make_fx()
    print(f"\n✅ All 7 templates created in templates/ folder")
