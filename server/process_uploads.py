"""
process_uploads.py — Cloud data processor
Reads uploaded Excel files, calculates territory P&L, pushes to Supabase.

Called by server: python3 process_uploads.py --period 2026-04 --upload-dir /tmp/uploads/session_id
Outputs JSON progress lines for streaming to frontend.
"""
import sys, os, json, argparse
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

def emit(log, pct=None, done=False, error=None):
    d = {}
    if log:   d['log'] = log
    if pct:   d['pct'] = pct
    if done:  d['done'] = True
    if error: d['error'] = error
    print(json.dumps(d), flush=True)

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    emit(None, error='openpyxl not installed. Run: pip install openpyxl')
    sys.exit(1)

# ── Parse args ────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument('--period', required=True)          # e.g. 2026-04
parser.add_argument('--upload-dir', required=True)      # dir with uploaded files
parser.add_argument('--client-id', default=None)        # for multi-tenant
args = parser.parse_args()

PERIOD     = args.period
UPLOAD_DIR = args.upload_dir
CLIENT_ID  = args.client_id

emit(f'Processing period: {PERIOD}', 5)
emit(f'Upload dir: {UPLOAD_DIR}', 6)

# ── FX Rates ──────────────────────────────────────────────────────────────────
DEFAULT_FX = {
    'MYR':1.0,'KRW':0.002654,'JPY':0.02522,'EUR':4.619,'USD':4.447,
    'AED':1.098,'MXN':0.2215,'BRL':0.7675,'AUD':2.759,'INR':0.04274,
    'IDR':0.0002374,'PHP':0.07748,'THB':0.1227,'SGD':3.29,'GBP':5.62,
    'HKD':0.571,'CNY':0.611,'TWD':0.136,'CHF':4.98,
}
fx = dict(DEFAULT_FX)

def find_file(prefix):
    """Find uploaded file matching prefix (case-insensitive)."""
    for f in os.listdir(UPLOAD_DIR):
        if f.lower().startswith(prefix.lower()) and f.lower().endswith('.xlsx'):
            return os.path.join(UPLOAD_DIR, f)
    return None

# ── Load FX Rates ─────────────────────────────────────────────────────────────
fx_file = find_file('fx_rates')
if fx_file:
    emit('Loading FX rates...', 8)
    try:
        wb = load_workbook(fx_file, data_only=True, read_only=True)
        ws = wb.active
        # Smart header detection: find row with 'Currency' header
        fx_data_row = 2
        for try_row in [1, 2, 3, 4, 5]:
            test = [str(c or '').strip().lower() for c in next(ws.iter_rows(min_row=try_row, max_row=try_row, values_only=True))]
            if 'currency' in test:
                # Find column positions
                ci_cur = test.index('currency')
                ci_rate = next((i for i,h in enumerate(test) if 'rate' in h), ci_cur + 1)
                fx_data_row = try_row + 1
                break
        else:
            ci_cur, ci_rate, fx_data_row = 0, 1, 2

        for row in ws.iter_rows(min_row=fx_data_row, values_only=True):
            if row and len(row) > ci_rate and row[ci_cur] and row[ci_rate]:
                try:
                    ccy = str(row[ci_cur]).strip().upper()
                    rate = float(row[ci_rate])
                    if ccy and rate > 0:
                        fx[ccy] = rate
                except: pass
        wb.close()
        emit(f'FX rates loaded: {len(fx)} currencies', 10)
    except Exception as e:
        emit(f'FX load warning: {e}')

def to_myr(amount, currency):
    try: return float(amount or 0) * fx.get(str(currency).upper(), 1.0)
    except: return 0.0

def safe(v, d=0.0):
    try: return float(v) if v not in (None, '') else d
    except: return d

# ── Territory detection from region/country ───────────────────────────────────
REGION_MAP = {
    'korea':'Korea','south korea':'Korea','kr':'Korea',
    'japan':'Japan','jp':'Japan',
    'europe':'Europe','eu':'Europe','de':'Europe','fr':'Europe','gb':'Europe',
    'uk':'Europe','nl':'Europe','be':'Europe','at':'Europe','ch':'Europe',
    'gcc':'GCC','ae':'GCC','sa':'GCC','kw':'GCC','qa':'GCC','bh':'GCC','om':'GCC',
    'uae':'GCC','saudi':'GCC',
    'usa':'USA','us':'USA','united states':'USA',
    'latam':'Latam','mexico':'Latam','mx':'Latam','colombia':'Latam','co':'Latam',
    'chile':'Latam','cl':'Latam','argentina':'Latam','ar':'Latam','peru':'Latam','pe':'Latam',
    'brasil':'Brasil','brazil':'Brasil','br':'Brasil',
    'oceania':'Oceania','australia':'Oceania','au':'Oceania','new zealand':'Oceania','nz':'Oceania',
    'india':'India','in':'India',
    'indonesia':'Indonesia','id':'Indonesia',
    'philippines':'Philippines','ph':'Philippines',
    'thailand':'Thailand','th':'Thailand',
    'malaysia':'Malaysia','my':'Malaysia',
    'molnu':'Molnu',
}
ALL_TERRITORIES = ['Korea','Japan','Europe','GCC','USA','Latam','Brasil','Oceania',
                   'India','Indonesia','Philippines','Thailand','Malaysia','Molnu']

def detect_territory(region_str, country_str=''):
    for s in [region_str, country_str]:
        if s:
            key = str(s).lower().strip()
            if key in REGION_MAP: return REGION_MAP[key]
            for k,v in REGION_MAP.items():
                if k in key: return v
    return 'Malaysia'  # default

# ── Init territory accumulators ───────────────────────────────────────────────
def empty_territory(name):
    return {
        'territory':name,'period':PERIOD,'client_id':CLIENT_ID,
        'brand':'','currency':'MYR','local_currency':'MYR','fx_rate_to_myr':1.0,
        'gross':0,'net':0,'refund_total':0,'discount':0,'tax':0,'shipping':0,
        'orders':0,'orders_paid':0,'orders_unpaid':0,'orders_refunded':0,
        'fee_total':0,'fee_stripe':0,'fee_paypal':0,'fee_payex':0,
        'fee_tiktok':0,'fee_shopee':0,'fee_lazada':0,'fee_xendit':0,
        'gw_payex':0,'gw_stripe_gross':0,'gw_paypal_gross':0,'gw_xendit_gross':0,
        'gw_settlement_net':0,'payment':0,'dbt':0,'aov':0,'margin_pct':0,
        'fulfilled':0,'unfulfilled':0,'cogs':0,'gross_profit':0,
        'payment_methods':{},'daily':{},'products':[],
    }

data = {t: empty_territory(t) for t in ALL_TERRITORIES}

# ── Helper: find multiple files matching a prefix ─────────────────────────────
def find_files(prefix):
    """Find ALL uploaded files matching prefix (handles multiple territory files)."""
    found = []
    for f in os.listdir(UPLOAD_DIR):
        flower = f.lower()
        if flower.startswith(prefix.lower()) and (flower.endswith('.xlsx') or flower.endswith('.csv') or flower.endswith('.json')):
            found.append(os.path.join(UPLOAD_DIR, f))
    return found

def read_csv_orders(filepath, territory_override=None):
    """Read orders from a CSV file (per-territory export format)."""
    import csv as csv_module
    rows_data = []
    try:
        with open(filepath, encoding='utf-8-sig', errors='replace') as fh:
            reader = csv_module.DictReader(fh)
            for row in reader:
                rows_data.append((row, territory_override))
    except Exception as e:
        emit(f'CSV read warning {filepath}: {e}')
    return rows_data

def process_order_row(row_dict, territory_override=None):
    """Process one order row dict → accumulate into data dict."""
    # Normalise keys to lowercase
    row = {k.lower().strip(): v for k, v in row_dict.items()}
    cur = str(row.get('currency', 'MYR') or 'MYR').strip()
    region = str(row.get('region', '') or '')
    billing_country = str(row.get('billing country', '') or row.get('delivery country', ''))
    ter = territory_override or detect_territory(region, billing_country)
    d = data[ter]

    gross_native = safe(row.get('total', 0))
    ref_native   = safe(row.get('refunded amount', 0))
    ship_native  = safe(row.get('shipping rate', 0))
    tax_native   = safe(row.get('total tax', 0))
    net_native   = gross_native - ref_native

    d['gross']        += to_myr(gross_native, cur)
    d['net']          += to_myr(net_native, cur)
    d['refund_total'] += to_myr(ref_native, cur)
    d['shipping']     += to_myr(ship_native, cur)
    d['tax']          += to_myr(tax_native, cur)
    d['orders']       += 1

    pay = str(row.get('payment status', '') or '').strip()
    ful = str(row.get('fulfillment status', '') or '').strip()
    if pay == 'Paid': d['orders_paid'] += 1
    elif pay in ('Refunded','Partially refunded'): d['orders_refunded'] += 1
    else: d['orders_unpaid'] += 1
    if ful == 'Fulfilled': d['fulfilled'] += 1
    elif ful == 'Unfulfilled': d['unfulfilled'] += 1

# ── Process Wix Orders ────────────────────────────────────────────────────────
# Handle 3 formats:
# 1. Single combined XLSX (March format: has 'Region' col)
# 2. Per-territory CSVs (April format: one CSV per territory)
# 3. API-fetched JSON (from Wix API with territory already tagged)

# Check for per-territory CSVs first (April format: files named by territory)
TERRITORY_CSV_MAP = {
    'korea': 'Korea', 'japan': 'Japan', 'europe': 'Europe', 'gcc': 'GCC',
    'usa': 'USA', 'latam': 'Latam', 'brazil': 'Brasil', 'brasil': 'Brasil',
    'oceania': 'Oceania', 'india': 'India', 'indonesia': 'Indonesia',
    'philippines': 'Philippines', 'thailand': 'Thailand', 'malaysia': 'Malaysia',
    'molnu': 'Molnu',
}
ter_csvs = []
for f in os.listdir(UPLOAD_DIR):
    fl = f.lower()
    if fl.endswith('.csv'):
        for key, ter in TERRITORY_CSV_MAP.items():
            if key in fl and ('order' in fl or 'payment' not in fl):
                ter_csvs.append((os.path.join(UPLOAD_DIR, f), ter))
                break

if ter_csvs:
    emit(f'Reading per-territory order CSVs ({len(ter_csvs)} files)...', 15)
    total_rows = 0
    for fpath, territory in ter_csvs:
        rows = read_csv_orders(fpath, territory)
        for row_dict, ter_override in rows:
            process_order_row(row_dict, ter_override)
            total_rows += 1
    emit(f'Orders (CSV): {total_rows} rows across territories', 30)

# Also check for API JSON format
api_json = find_file('wix_orders')
if api_json and api_json.endswith('.json'):
    emit('Reading Wix Orders from API fetch...', 15)
    import json as _json
    with open(api_json, encoding='utf-8') as f:
        api_data = _json.load(f)
    for order in api_data.get('rows', []):
        ter = order.get('_territory', 'Malaysia')
        cur = order.get('currency', 'MYR')
        gross = safe(order.get('priceSummary', {}).get('total', {}).get('amount', 0))
        refund = safe(order.get('priceSummary', {}).get('refund', {}).get('amount', 0))
        d = data.get(ter, data['Malaysia'])
        d['gross'] += to_myr(gross, cur)
        d['net']   += to_myr(gross - refund, cur)
        d['orders'] += 1
        if order.get('paymentStatus') == 'PAID': d['orders_paid'] += 1
    emit(f'Orders (API): {len(api_data.get("rows",[]))} from Wix API', 30)

orders_file = find_file('wix_orders')
if orders_file and orders_file.endswith('.xlsx') and not ter_csvs:
    emit('Reading Wix Orders...', 15)
    try:
        wb = load_workbook(orders_file, data_only=True, read_only=True)
        ws = wb.active
        # Smart header detection: find the row with 'Region' header (may be row 1 or 2)
        header_row = 1
        data_start_row = 2
        for try_row in [1, 2, 3, 4]:
            test_hdrs = [str(c or '').strip().lower() for c in next(ws.iter_rows(min_row=try_row, max_row=try_row, values_only=True))]
            if 'region' in test_hdrs:
                header_row = try_row
                data_start_row = try_row + 1
                break
        headers_raw = [str(c or '').strip() for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        # Normalise: strip case for matching
        headers = [h.lower() for h in headers_raw]
        def col(name):
            n = name.lower()
            if n in headers: return headers.index(n)
            # Partial match fallback
            for i, h in enumerate(headers):
                if n in h or h in n: return i
            return -1

        ci_region   = col('Region')
        ci_currency = col('Currency')
        ci_total    = col('Total')
        ci_net      = col('Net amount')        # Wix field: Total - Refunds
        ci_refund   = col('Refunded amount')
        ci_ship     = col('Shipping rate')
        ci_tax      = col('Total tax')
        ci_discount = col('Discount')
        ci_coupon   = col('Coupon code')
        ci_paystatus= col('Payment status')
        ci_fulstatus= col('Fulfillment status')
        ci_country  = col('Billing country')

        rows = 0
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not row or not any(row): continue
            # Skip sample/note rows
            r0 = str(row[0] or '').strip()
            if r0.startswith('ℹ') or r0.lower() in ('region','sample','example'): continue

            cur = str(row[ci_currency] if ci_currency >= 0 and len(row) > ci_currency else 'MYR' or 'MYR').strip()
            region = str(row[ci_region] if ci_region >= 0 and len(row) > ci_region else '' or '').strip()
            country = str(row[ci_country] if ci_country >= 0 and len(row) > ci_country else '' or '').strip()
            ter = detect_territory(region, country)
            d = data[ter]

            gross_native = safe(row[ci_total] if ci_total >= 0 and len(row) > ci_total else 0)
            ref_native   = safe(row[ci_refund] if ci_refund >= 0 and len(row) > ci_refund else 0)
            ship_native  = safe(row[ci_ship]  if ci_ship >= 0 and len(row) > ci_ship else 0)
            tax_native   = safe(row[ci_tax]   if ci_tax >= 0 and len(row) > ci_tax else 0)
            disc_native  = safe(row[ci_discount] if ci_discount >= 0 and len(row) > ci_discount else 0)
            # net = Total - Refunds (Wix 'Net amount' also deducts gateway fees - don't use it)
            net_native   = gross_native - ref_native

            d['gross']        += to_myr(gross_native, cur)
            d['net']          += to_myr(net_native, cur)
            d['refund_total'] += to_myr(ref_native, cur)
            d['shipping']     += to_myr(ship_native, cur)
            d['tax']          += to_myr(tax_native, cur)
            d['discount']     += to_myr(disc_native, cur)  # coupon is text field, not numeric
            d['orders']       += 1

            pay_status = str(row[ci_paystatus] if ci_paystatus >= 0 and len(row) > ci_paystatus else '' or '').strip()
            ful_status = str(row[ci_fulstatus] if ci_fulstatus >= 0 and len(row) > ci_fulstatus else '' or '').strip()
            if pay_status == 'Paid': d['orders_paid'] += 1
            elif pay_status in ('Refunded','Partially refunded'): d['orders_refunded'] += 1
            else: d['orders_unpaid'] += 1
            if ful_status == 'Fulfilled': d['fulfilled'] += 1
            elif ful_status == 'Unfulfilled': d['unfulfilled'] += 1
            rows += 1

        wb.close()
        total_orders = sum(d['orders'] for d in data.values())
        emit(f'Wix Orders: {rows} rows → {total_orders} orders across territories', 30)
    except Exception as e:
        emit(f'Wix Orders warning: {e}')

# ── Process Gateway Reports ───────────────────────────────────────────────────
# Stripe
stripe_file = find_file('stripe')
if stripe_file:
    emit('Reading Stripe report...', 45)
    try:
        wb = load_workbook(stripe_file, data_only=True, read_only=True)
        # Try "Stripe report 2" sheet first, else active
        ws = wb['Stripe report 2'] if 'Stripe report 2' in wb.sheetnames else wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1,max_row=1))]
        col_amount = next((i for i,h in enumerate(headers) if 'amount' in h.lower() and 'myr' in h.lower()), 3)
        col_fee    = next((i for i,h in enumerate(headers) if 'fee' in h.lower() and 'myr' in h.lower()), 4)
        col_type   = next((i for i,h in enumerate(headers) if 'type' in h.lower()), 1)
        col_country= next((i for i,h in enumerate(headers) if 'country' in h.lower()), -1)

        total_fee = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            type_val = str(row[col_type] if len(row) > col_type else '' or '').strip().lower()
            if type_val != 'charge': continue
            amount = safe(row[col_amount] if len(row) > col_amount else 0)
            fee    = safe(row[col_fee] if len(row) > col_fee else 0)
            country = str(row[col_country] if col_country >= 0 and len(row) > col_country else '' or '')
            ter = detect_territory('', country)
            data[ter]['gw_stripe_gross'] += amount
            data[ter]['fee_stripe']      += fee
            total_fee += fee
        wb.close()
        emit(f'Stripe: total fee = MYR {total_fee:,.2f}', 50)
    except Exception as e:
        emit(f'Stripe warning: {e}')

# PayPal
paypal_file = find_file('paypal')
if paypal_file:
    emit('Reading PayPal report...', 55)
    try:
        wb = load_workbook(paypal_file, data_only=True, read_only=True)
        ws = wb['Paypal'] if 'Paypal' in wb.sheetnames else wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1,max_row=1))]
        col_type   = next((i for i,h in enumerate(headers) if h.lower() in ('type','description')), 3)
        col_status = next((i for i,h in enumerate(headers) if 'balance impact' in h.lower()), 23)
        col_gross  = next((i for i,h in enumerate(headers) if h.lower() == 'gross'), 5)
        col_fee    = next((i for i,h in enumerate(headers) if h.lower() == 'fee'), 6)
        col_cur    = next((i for i,h in enumerate(headers) if h.lower() == 'currency'), 4)
        col_country= next((i for i,h in enumerate(headers) if 'country' in h.lower()), 22)

        total_fee = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            type_val = str(row[col_type] if len(row) > col_type else '' or '')
            status   = str(row[col_status] if len(row) > col_status else '' or '')
            if 'express checkout' not in type_val.lower() and 'payment' not in type_val.lower(): continue
            if 'debit' not in status.lower() and 'completed' not in status.lower(): continue
            cur    = str(row[col_cur] if len(row) > col_cur else 'USD' or 'USD').strip()
            gross  = safe(row[col_gross] if len(row) > col_gross else 0)
            fee    = abs(safe(row[col_fee] if len(row) > col_fee else 0))
            country= str(row[col_country] if col_country >= 0 and len(row) > col_country else '' or '')
            ter = detect_territory('', country)
            data[ter]['gw_paypal_gross'] += to_myr(gross, cur)
            data[ter]['fee_paypal']      += to_myr(fee, cur)
            total_fee += to_myr(fee, cur)
        wb.close()
        emit(f'PayPal: total fee = MYR {total_fee:,.2f}', 60)
    except Exception as e:
        emit(f'PayPal warning: {e}')

# Payex
payex_file = find_file('payex')
if payex_file:
    emit('Reading Payex report...', 65)
    try:
        wb = load_workbook(payex_file, data_only=True, read_only=True)
        ws = wb['Payex Report 2'] if 'Payex Report 2' in wb.sheetnames else wb.active
        headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1,max_row=1))]
        # Actual Payex Report 2 columns: Type(7), Gross(20), MDR(21), Net(23)
        col_type  = next((i for i,h in enumerate(headers) if h == 'type'), 7)
        col_gross = next((i for i,h in enumerate(headers) if h == 'gross'), 20)
        col_mdr   = next((i for i,h in enumerate(headers) if h == 'mdr'), 21)
        col_net   = next((i for i,h in enumerate(headers) if h == 'net'), 23)

        col_currency = next((i for i,h in enumerate(headers) if h == 'currency'), 19)
        total_mdr = 0; total_gross = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            type_val = str(row[col_type] if len(row) > col_type else '' or '').lower()
            if 'settlement' not in type_val: continue
            cur    = str(row[col_currency] if len(row) > col_currency else 'MYR' or 'MYR').strip()
            gross  = safe(row[col_gross] if len(row) > col_gross else 0)
            mdr    = safe(row[col_mdr] if len(row) > col_mdr else 0)
            gross_myr = to_myr(gross, cur)
            mdr_myr   = to_myr(mdr, cur)
            # Territory from currency (rough mapping — proper join done in reconciliation)
            ter = {'MYR':'Malaysia','KRW':'Korea','JPY':'Japan','IDR':'Indonesia',
                   'PHP':'Philippines','THB':'Thailand','BRL':'Brasil',
                   'AED':'GCC','SAR':'GCC','USD':'USA','EUR':'Europe',
                   'AUD':'Oceania','INR':'India','MXN':'Latam'}.get(cur, 'Malaysia')
            data[ter]['gw_payex']  += gross_myr
            data[ter]['fee_payex'] += mdr_myr
            total_mdr += mdr_myr; total_gross += gross_myr
        wb.close()
        emit(f'Payex: gross = MYR {total_gross:,.2f} | MDR = MYR {total_mdr:,.2f}', 70)
    except Exception as e:
        emit(f'Payex warning: {e}')

# Marketplace
mkt_file = find_file('marketplace')
if mkt_file:
    emit('Reading Marketplace report...', 75)
    try:
        wb = load_workbook(mkt_file, data_only=True, read_only=True)
        for sheet_name, fee_key in [('TikTok','fee_tiktok'),('Shopee','fee_shopee'),('Lazada','fee_lazada')]:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1,max_row=1))]
            col_fee    = next((i for i,h in enumerate(headers) if 'fee' in h.lower() or 'commission' in h.lower()), -1)
            col_region = next((i for i,h in enumerate(headers) if 'region' in h.lower()), -1)
            col_cur    = next((i for i,h in enumerate(headers) if 'currency' in h.lower()), -1)
            if col_fee < 0: continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row): continue
                fee = safe(row[col_fee] if len(row) > col_fee else 0)
                cur = str(row[col_cur] if col_cur >= 0 and len(row) > col_cur else 'MYR' or 'MYR').strip()
                region = str(row[col_region] if col_region >= 0 and len(row) > col_region else 'Malaysia' or 'Malaysia').strip()
                ter = detect_territory(region)
                data[ter][fee_key] += to_myr(fee, cur)
        wb.close()
        emit(f'Marketplace fees loaded', 78)
    except Exception as e:
        emit(f'Marketplace warning: {e}')

# ── Finalize calculations ─────────────────────────────────────────────────────
emit('Finalizing territory P&L...', 82)
for ter, d in data.items():
    if d['orders'] == 0 and d['gross'] == 0: continue  # skip empty territories
    d['fee_total'] = (d['fee_payex'] + d['fee_stripe'] + d['fee_paypal'] +
                      d['fee_tiktok'] + d['fee_shopee'] + d['fee_lazada'] + d['fee_xendit'])
    d['payment']   = (d['gw_payex'] + d['gw_stripe_gross'] + d['gw_paypal_gross'] + d['gw_xendit_gross'])
    d['gw_settlement_net'] = d['payment'] - d['fee_total']
    d['gross_profit'] = d['net'] - d['fee_total'] - d['cogs']
    if d['gross'] > 0:
        d['margin_pct'] = round(d['gross_profit'] / d['gross'] * 100, 2)
    if d['orders'] > 0:
        d['aov'] = round(d['net'] / d['orders'], 2)
    # Round all monetary fields
    for k in ['gross','net','refund_total','discount','tax','shipping','fee_total',
              'fee_stripe','fee_paypal','fee_payex','fee_tiktok','fee_shopee','fee_lazada',
              'gw_payex','gw_stripe_gross','gw_paypal_gross','payment','gw_settlement_net','gross_profit']:
        d[k] = round(d[k], 2)

# ── Push to Supabase ──────────────────────────────────────────────────────────
emit('Pushing to Supabase database...', 88)

supabase_url = os.environ.get('SUPABASE_URL','')
supabase_key = os.environ.get('SUPABASE_SERVICE_KEY','')

pushed = 0
errors = 0

if supabase_url and supabase_key:
    import urllib.request, urllib.error
    for ter, d in data.items():
        if d['orders'] == 0 and d['gross'] == 0: continue
        d['data_source'] = 'manual_upload'   # tag uploaded data separately from system workbooks
        payload = json.dumps(d).encode('utf-8')
        req = urllib.request.Request(
            f"{supabase_url}/rest/v1/territory_data?on_conflict=territory,period,data_source",
            data=payload,
            headers={
                'apikey': supabase_key,
                'Authorization': f'Bearer {supabase_key}',
                'Content-Type': 'application/json',
                'Prefer': 'resolution=merge-duplicates,return=minimal',
            },
            method='POST'
        )
        try:
            urllib.request.urlopen(req, timeout=15)
            pushed += 1
            emit(f'  ✓ {ter}: Net MYR {d["net"]:,.0f} | Orders {d["orders"]}')
        except urllib.error.HTTPError as e:
            body = e.read().decode('utf-8', errors='replace')[:200]
            emit(f'  ⚠ {ter}: {e.code} {body}')
            errors += 1

    emit(f'Supabase: {pushed} territories saved ({errors} errors)', 95)
else:
    emit('⚠ Supabase not configured — saving to local cache only')

# ── Summary ───────────────────────────────────────────────────────────────────
total_net    = sum(d['net'] for d in data.values())
total_orders = sum(d['orders'] for d in data.values())
emit(f'Total Net Revenue: MYR {total_net:,.2f} | Total Orders: {total_orders}', 98)
emit(f'Build complete!', 100, done=True)
